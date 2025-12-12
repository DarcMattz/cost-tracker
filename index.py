import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import openpyxl
import os
import pyperclip
from datetime import date

EXCEL_FILE = "cost_data.xlsx"

# Column Structure (12 columns)
HEADERS = [
    "Item Description","Category","Unit","Currency","Bare Price",
    "Brand","Supplier","Location","Date",
    "Logged by","Reference","Remarks"
]

# Create Excel file if missing
if not os.path.exists(EXCEL_FILE):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CostData"
    ws.append(HEADERS)
    wb.save(EXCEL_FILE)

# Load Excel Data
def load_data():
    wb = openpyxl.load_workbook(EXCEL_FILE)
    ws = wb.active
    return [row for row in ws.iter_rows(min_row=2, values_only=True)]

# Save Excel Data
def save_data(data):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "CostData"

    ws.append(HEADERS)
    for row in data:
        ws.append(list(row))

    wb.save(EXCEL_FILE)

# Add/Edit Item
def save_item():
    global data, edit_index

    item_desc = item_entry.get().strip()
    category = category_entry.get().strip()
    unit = unit_entry.get().strip()
    currency = currency_entry.get().strip()

    try:
        raw_cost = material_entry.get().replace(",", "")
        bare_price = float(raw_cost or 0)
        bare_price_formatted = "{:,.2f}".format(bare_price)
    except ValueError:
        messagebox.showerror("Error", "Bare Price must be numeric.")
        return

    brand = brand_entry.get().strip()
    supplier = supplier_entry.get().strip()
    location = location_entry.get().strip()
    item_date = date_entry.get().strip() or str(date.today())
    logged_by = logged_entry.get().strip()
    reference_path = reference_var.get().strip()
    remarks = notes_entry.get().strip()

    if not item_desc:
        messagebox.showwarning("Input Error", "Item Description is required.")
        return

    new_row = (
        item_desc, category, unit, currency, bare_price_formatted,
        brand, supplier, location, item_date,
        logged_by, reference_path, remarks
    )

    if edit_index is not None:
        data[edit_index] = new_row
        edit_index = None
    else:
        data.append(new_row)

    save_data(data)
    render_table()
    clear_form()

# Delete Item
def delete_item():
    global data
    selected = tree.selection()
    if not selected:
        messagebox.showwarning("Select", "Select an item to delete.")
        return

    idx = int(selected[0])
    if messagebox.askyesno("Confirm", "Delete this item?"):
        data.pop(idx)
        save_data(data)
        render_table()

# Edit Item
def edit_item():
    global edit_index
    selected = tree.selection()
    if not selected:
        messagebox.showwarning("Select", "Select an item to edit.")
        return

    idx = int(selected[0])
    row = data[idx]
    edit_index = idx

    item_entry.delete(0, tk.END)
    item_entry.insert(0, row[0])

    category_entry.delete(0, tk.END)
    category_entry.insert(0, row[1])

    unit_entry.delete(0, tk.END)
    unit_entry.insert(0, row[2])

    currency_entry.delete(0, tk.END)
    currency_entry.insert(0, row[3])

    material_entry.delete(0, tk.END)
    material_entry.insert(0, row[4].replace(",", ""))

    brand_entry.delete(0, tk.END)
    brand_entry.insert(0, row[5])

    supplier_entry.delete(0, tk.END)
    supplier_entry.insert(0, row[6])

    location_entry.delete(0, tk.END)
    location_entry.insert(0, row[7])

    date_entry.delete(0, tk.END)
    date_entry.insert(0, row[8])

    logged_entry.delete(0, tk.END)
    logged_entry.insert(0, row[9])

    reference_var.set(row[10] or "")

    notes_entry.delete(0, tk.END)
    notes_entry.insert(0, row[11])

# Clear Form
def clear_form():
    global edit_index
    edit_index = None

    item_entry.delete(0, tk.END)
    category_entry.delete(0, tk.END)
    unit_entry.delete(0, tk.END)
    currency_entry.delete(0, tk.END)
    material_entry.delete(0, tk.END)
    brand_entry.delete(0, tk.END)
    supplier_entry.delete(0, tk.END)
    location_entry.delete(0, tk.END)

    date_entry.delete(0, tk.END)
    date_entry.insert(0, str(date.today()))

    logged_entry.delete(0, tk.END)
    reference_var.set("")
    notes_entry.delete(0, tk.END)

# Browse Reference
def browse_reference():
    path = filedialog.askopenfilename()
    if path:
        reference_var.set(path)

# Right-click: Open Reference
def open_reference():
    selected = tree.selection()
    if not selected:
        return
    idx = int(selected[0])
    ref = data[idx][10]
    if not ref:
        messagebox.showinfo("No Reference", "No reference assigned.")
        return
    if not os.path.exists(ref):
        messagebox.showerror("Error", "Referenced file does not exist.")
        return
    os.startfile(ref)

# Copy Reference
def copy_reference():
    selected = tree.selection()
    if selected:
        pyperclip.copy(data[int(selected[0])][10] or "")

# Paste Reference
def paste_reference():
    global data
    selected = tree.selection()
    if not selected:
        return
    idx = int(selected[0])
    paste_val = pyperclip.paste()
    row = list(data[idx])
    row[10] = paste_val
    data[idx] = tuple(row)
    save_data(data)
    render_table()

# Right-click Menu
def show_context_menu(event):
    row = tree.identify_row(event.y)
    if row:
        tree.selection_set(row)
        menu.post(event.x_root, event.y_root)

# Render Table
def render_table():
    query = search_var.get().lower()
    for i in tree.get_children():
        tree.delete(i)
    for idx, row in enumerate(data):
        if query in str(row).lower():
            tree.insert("", "end", iid=idx, values=row)

# ================= GUI SETUP =================
root = tk.Tk()
root.title("Cost Database")
root.configure(bg="#f0f2f5")

try:
    root.state("zoomed")
except:
    root.geometry("1500x850")

# Search Bar
search_frame = tk.Frame(root, bg="#f0f2f5")
search_frame.pack(fill=tk.X, padx=10, pady=5)

search_var = tk.StringVar()
tk.Label(search_frame, text="Search:", font=("Arial", 12), bg="#f0f2f5").pack(side=tk.LEFT)
search_entry = tk.Entry(search_frame, textvariable=search_var, font=("Arial", 12), width=40)
search_entry.pack(side=tk.LEFT, padx=10)
search_entry.bind("<KeyRelease>", lambda e: render_table())

# Table
table_frame = tk.Frame(root)
table_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=5)

columns = HEADERS
tree = ttk.Treeview(table_frame, columns=columns, show="headings")

# Set all column widths equal
column_width = 120
for col in columns:
    tree.heading(col, text=col)
    tree.column(col, width=column_width)

tree.pack(fill=tk.BOTH, expand=True, side=tk.LEFT)

# Scrollbar
scrollbar = ttk.Scrollbar(table_frame, orient="vertical", command=tree.yview)
tree.configure(yscroll=scrollbar.set)
scrollbar.pack(side=tk.RIGHT, fill=tk.Y)

# Right-click Menu
menu = tk.Menu(root, tearoff=0)
menu.add_command(label="Open Reference", command=open_reference)
menu.add_command(label="Copy Reference", command=copy_reference)
menu.add_command(label="Paste Reference", command=paste_reference)
tree.bind("<Button-3>", show_context_menu)

# Form Section
form_frame = tk.LabelFrame(root, text="Add / Edit Item", padx=10, pady=10)
form_frame.pack(fill=tk.X, padx=10, pady=10)

tk.Label(form_frame, text="Item Description").grid(row=0, column=0)
item_entry = tk.Entry(form_frame)
item_entry.grid(row=0, column=1)

tk.Label(form_frame, text="Category").grid(row=0, column=2)
category_entry = tk.Entry(form_frame)
category_entry.grid(row=0, column=3)

tk.Label(form_frame, text="Unit").grid(row=1, column=0)
unit_entry = tk.Entry(form_frame)
unit_entry.grid(row=1, column=1)

tk.Label(form_frame, text="Currency").grid(row=1, column=2)
currency_entry = tk.Entry(form_frame)
currency_entry.grid(row=1, column=3)

tk.Label(form_frame, text="Bare Price").grid(row=2, column=0)
material_entry = tk.Entry(form_frame)
material_entry.grid(row=2, column=1)

tk.Label(form_frame, text="Brand").grid(row=2, column=2)
brand_entry = tk.Entry(form_frame)
brand_entry.grid(row=2, column=3)

tk.Label(form_frame, text="Supplier").grid(row=3, column=0)
supplier_entry = tk.Entry(form_frame)
supplier_entry.grid(row=3, column=1)

tk.Label(form_frame, text="Location").grid(row=3, column=2)
location_entry = tk.Entry(form_frame)
location_entry.grid(row=3, column=3)

tk.Label(form_frame, text="Date").grid(row=4, column=0)
date_entry = tk.Entry(form_frame)
date_entry.grid(row=4, column=1)
date_entry.insert(0, str(date.today()))

tk.Label(form_frame, text="Logged by").grid(row=4, column=2)
logged_entry = tk.Entry(form_frame)
logged_entry.grid(row=4, column=3)

tk.Label(form_frame, text="Reference").grid(row=5, column=0)
reference_var = tk.StringVar()
reference_entry = tk.Entry(form_frame, textvariable=reference_var, width=30)
reference_entry.grid(row=5, column=1)
tk.Button(form_frame, text="Browse", command=browse_reference).grid(row=5, column=2, padx=5)

tk.Label(form_frame, text="Remarks").grid(row=5, column=3)
notes_entry = tk.Entry(form_frame, width=40)
notes_entry.grid(row=5, column=4, columnspan=3)

btn_frame = tk.Frame(form_frame)
btn_frame.grid(row=6, column=0, columnspan=6, pady=10)

tk.Button(btn_frame, text="Save", bg="#4CAF50", fg="white", width=12, command=save_item).pack(side=tk.LEFT, padx=5)
tk.Button(btn_frame, text="Clear", bg="#f0ad4e", fg="white", width=12, command=clear_form).pack(side=tk.LEFT, padx=5)
tk.Button(btn_frame, text="Edit Selected", bg="#2196F3", fg="white", width=12, command=edit_item).pack(side=tk.LEFT, padx=5)
tk.Button(btn_frame, text="Delete Selected", bg="#f44336", fg="white", width=12, command=delete_item).pack(side=tk.LEFT, padx=5)

# Watermark
watermark = tk.Label(root, text="Jibee | VCC", font=("Arial", 10, "italic"),
                     bg="#f0f2f5", fg="#999999")
watermark.place(relx=1.0, rely=1.0, anchor="se", x=-10, y=-10)

# Load & Start
data = load_data()
edit_index = None
render_table()

root.mainloop()
